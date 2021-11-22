#Ver 1.1.1

import Weather
import openpyxl
import time
from notify_run import Notify
from importlib import reload
notify = Notify()
wb = openpyxl.load_workbook("Weather.xlsx")
sheet = wb['Sheet1']



#Fixes glitch which displays the incorrect number of rows
def maximunrows(sheet):
    Falsevalue = sheet.max_row
    for i in range(sheet.max_row):
        if sheet.cell(row=i+1, column=1).value == None:
            Falsevalue -= 1
    return Falsevalue

def update(x,y):
    if x == y:
        return False
    else:
        return True


def checkupdateall():
    global currhigh, currlow, currrainfall, currdir, currspeed, currsky
    if update(currhigh, Weather.high) or update(currlow,Weather.low) or update(currrainfall, Weather.rainfall) or update(currdir, Weather.winddirection) or update(currspeed, Weather.windspeed) or update(currsky, Weather.sky):
        return True
    else:
        return False


#Find Whats the Latest Information so the program is able to check weather to update new or not
maxcurrentrows = maximunrows(sheet)
currhigh = sheet.cell(row=maxcurrentrows, column=3).value
currlow = sheet.cell(row=maxcurrentrows, column=4).value
currrainfall = sheet.cell(row=maxcurrentrows, column=5).value
currdir = sheet.cell(row=maxcurrentrows, column=6).value
currspeed = sheet.cell(row=maxcurrentrows, column=7).value
currsky = sheet.cell(row=maxcurrentrows, column=8).value

#Mainloop
while True:
    reload(Weather)
    #Row to write on
    nowrow = maximunrows(sheet)+1
    #Check if the information has changed and if so, write new cells
    if checkupdateall():
        sheet.cell(row=nowrow, column=1).value = Weather.datestamp
        sheet.cell(row=nowrow, column=2).value = Weather.timestamp
        sheet.cell(row=nowrow, column=3).value = Weather.high
        sheet.cell(row=nowrow, column=4).value = Weather.low
        sheet.cell(row=nowrow, column=5).value = Weather.rainfall
        sheet.cell(row=nowrow, column=6).value = Weather.winddirection
        sheet.cell(row=nowrow, column=7).value = Weather.windspeed
        sheet.cell(row=nowrow, column=8).value = Weather.sky
        wb.save('Weather.xlsx')
        print("Saved at " + Weather.timestamp)
        notify.send('There is a weather update')
    else:
        print("Not saved, No change in weather. Timestamp: "+Weather.timestamp)
    #Update Latest weather
    currhigh = Weather.high
    currlow = Weather.low
    currrainfall = Weather.rainfall
    currdir = Weather.winddirection
    currspeed = Weather.windspeed
    currsky = Weather.sky


    #Wait for next check
    for i in range(600):
        time.sleep(1)